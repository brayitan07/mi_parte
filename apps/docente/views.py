# apps/docente/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db.models import Avg
from django.utils import timezone

from .models import (
    Docente, AsignacionDocente, Curso, Materia,
    Estudiante, Calificacion, Tarea, Asistencia, Mensaje,
)

# ─────────────────────────────────────────
# UTILIDAD: obtener el docente del usuario
# ─────────────────────────────────────────

def _get_docente(request):
    """Devuelve el Docente ligado al usuario autenticado o None."""
    try:
        return request.user.docente
    except Docente.DoesNotExist:
        return None


# ─────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────

@login_required
def dashboard_docente(request):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    asignaciones = AsignacionDocente.objects.filter(docente=docente).select_related('materia', 'curso')
    tareas       = Tarea.objects.filter(asignacion__docente=docente)
    mensajes_no_leidos = Mensaje.objects.filter(docente=docente, leido=False, enviado_por='estudiante').count()

    context = {
        'docente':            docente,
        'asignaciones':       asignaciones,
        'total_tareas':       tareas.count(),
        'mensajes_no_leidos': mensajes_no_leidos,
    }
    return render(request, 'docente/dashboard_docente.html', context)


# ─────────────────────────────────────────
# CURSOS
# ─────────────────────────────────────────

@login_required
def cursos_docente(request):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    asignaciones = (
        AsignacionDocente.objects
        .filter(docente=docente)
        .select_related('curso', 'materia')
        .prefetch_related('curso__estudiantes')
    )

    context = {
        'docente':      docente,
        'asignaciones': asignaciones,
    }
    return render(request, 'docente/cursos_docente.html', context)


# ─────────────────────────────────────────
# TAREAS
# ─────────────────────────────────────────

@login_required
def tarea_docente(request):
    """Lista todas las tareas del docente."""
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    tareas = (
        Tarea.objects
        .filter(asignacion__docente=docente)
        .select_related('asignacion__materia', 'asignacion__curso')
        .order_by('-fecha_limite')
    )

    context = {
        'docente': docente,
        'tareas':  tareas,
    }
    return render(request, 'docente/tarea_docente.html', context)


@login_required
def crear_tarea(request):
    """Muestra el formulario y crea una nueva tarea."""
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    asignaciones = (
        AsignacionDocente.objects
        .filter(docente=docente)
        .select_related('materia', 'curso')
    )

    if request.method == 'POST':
        asignacion_id = request.POST.get('asignacion')
        titulo        = request.POST.get('titulo', '').strip()
        descripcion   = request.POST.get('descripcion', '').strip()
        fecha_limite  = request.POST.get('fecha_limite')

        # Validación básica
        if not (asignacion_id and titulo and fecha_limite):
            context = {
                'docente':      docente,
                'asignaciones': asignaciones,
                'error':        'Por favor completa todos los campos obligatorios.',
                'form':         request.POST,
            }
            return render(request, 'docente/crear_tarea.html', context)

        asignacion = get_object_or_404(AsignacionDocente, id=asignacion_id, docente=docente)

        # Contamos los estudiantes del curso para total_estudiantes
        total_est = Estudiante.objects.filter(curso=asignacion.curso).count()

        Tarea.objects.create(
            asignacion        = asignacion,
            titulo            = titulo,
            descripcion       = descripcion,
            fecha_limite      = fecha_limite,
            total_estudiantes = total_est,
        )
        return redirect('tarea_docente')

    context = {
        'docente':      docente,
        'asignaciones': asignaciones,
    }
    return render(request, 'docente/crear_tarea.html', context)


@login_required
def detalle_tarea(request, id):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    tarea = get_object_or_404(Tarea, id=id, asignacion__docente=docente)
    context = {
        'docente': docente,
        'tarea':   tarea,
    }
    return render(request, 'docente/detalle_tarea.html', context)


@login_required
def editar_tarea(request, id):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    tarea        = get_object_or_404(Tarea, id=id, asignacion__docente=docente)
    asignaciones = (
        AsignacionDocente.objects
        .filter(docente=docente)
        .select_related('materia', 'curso')
    )

    if request.method == 'POST':
        asignacion_id = request.POST.get('asignacion')
        titulo        = request.POST.get('titulo', '').strip()
        descripcion   = request.POST.get('descripcion', '').strip()
        fecha_limite  = request.POST.get('fecha_limite')
        estado        = request.POST.get('estado', 'pendiente')

        if not (asignacion_id and titulo and fecha_limite):
            context = {
                'docente':      docente,
                'tarea':        tarea,
                'asignaciones': asignaciones,
                'error':        'Por favor completa todos los campos obligatorios.',
            }
            return render(request, 'docente/editar_tarea.html', context)

        asignacion = get_object_or_404(AsignacionDocente, id=asignacion_id, docente=docente)

        tarea.asignacion   = asignacion
        tarea.titulo       = titulo
        tarea.descripcion  = descripcion
        tarea.fecha_limite = fecha_limite
        tarea.estado       = estado
        tarea.save()
        return redirect('tarea_docente')

    context = {
        'docente':      docente,
        'tarea':        tarea,
        'asignaciones': asignaciones,
    }
    return render(request, 'docente/editar_tarea.html', context)


@login_required
def eliminar_tarea(request, id):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    tarea = get_object_or_404(Tarea, id=id, asignacion__docente=docente)
    tarea.delete()
    return redirect('tarea_docente')


# ─────────────────────────────────────────
# CALIFICACIONES
# ─────────────────────────────────────────

@login_required
def calificaciones_docente(request):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    # Cursos y materias que tiene este docente (para los <select>)
    asignaciones = (
        AsignacionDocente.objects
        .filter(docente=docente)
        .select_related('materia', 'curso')
    )
    cursos   = Curso.objects.filter(asignaciones__docente=docente).distinct()
    materias = Materia.objects.filter(asignaciones__docente=docente).distinct()

    curso_sel   = None
    materia_sel = None
    estudiantes = []
    calificaciones_map = {}  # {estudiante.id: Calificacion}

    # Estadísticas globales del docente
    todas_cal = Calificacion.objects.filter(asignacion__docente=docente)
    total_notas     = todas_cal.count()
    promedios       = [c.promedio for c in todas_cal]
    promedio_general = round(sum(promedios) / len(promedios), 2) if promedios else 0
    en_riesgo        = sum(1 for p in promedios if p < 3.0)
    destacados       = sum(1 for p in promedios if p >= 4.5)

    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')

    if curso_id and materia_id:
        curso_sel   = get_object_or_404(Curso,   id=curso_id)
        materia_sel = get_object_or_404(Materia, id=materia_id)

        # Verificamos que el docente tenga esa asignación
        asignacion = AsignacionDocente.objects.filter(
            docente=docente, curso=curso_sel, materia=materia_sel
        ).first()

        if asignacion:
            estudiantes = Estudiante.objects.filter(curso=curso_sel)
            cals = Calificacion.objects.filter(asignacion=asignacion)
            calificaciones_map = {c.estudiante_id: c for c in cals}

    context = {
        'docente':           docente,
        'cursos':            cursos,
        'materias':          materias,
        'curso_sel':         curso_sel,
        'materia_sel':       materia_sel,
        'estudiantes':       estudiantes,
        'calificaciones_map': calificaciones_map,
        'total_notas':       total_notas,
        'promedio_general':  promedio_general,
        'en_riesgo':         en_riesgo,
        'destacados':        destacados,
    }
    return render(request, 'docente/calificaciones_docente.html', context)


@login_required
@require_POST
def guardar_nota(request):
    """
    Llamado via fetch desde calificaciones_docente.html.
    Recibe: estudiante_id, curso_id, materia_id, tarea, parcial, examen
    Crea o actualiza la Calificacion usando get_or_create sobre la AsignacionDocente.
    """
    docente = _get_docente(request)
    if not docente:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante_id = request.POST.get('estudiante_id')
    curso_id      = request.POST.get('curso_id')
    materia_id    = request.POST.get('materia_id')

    try:
        tarea_val   = float(request.POST.get('tarea',   0))
        parcial_val = float(request.POST.get('parcial', 0))
        examen_val  = float(request.POST.get('examen',  0))
    except ValueError:
        return JsonResponse({'error': 'Valores numéricos inválidos'}, status=400)

    # Validar rango colombiano 0–5
    for v in (tarea_val, parcial_val, examen_val):
        if not (0 <= v <= 5):
            return JsonResponse({'error': 'Las notas deben estar entre 0 y 5'}, status=400)

    estudiante = get_object_or_404(Estudiante, id=estudiante_id)
    asignacion = get_object_or_404(
        AsignacionDocente,
        docente  = docente,
        curso_id = curso_id,
        materia_id = materia_id,
    )

    cal, _ = Calificacion.objects.get_or_create(
        estudiante = estudiante,
        asignacion = asignacion,
    )
    cal.tarea   = tarea_val
    cal.parcial = parcial_val
    cal.examen  = examen_val
    cal.save()

    return JsonResponse({
        'ok':      True,
        'promedio': cal.promedio,
    })


@login_required
@require_POST
def eliminar_nota(request):
    docente = _get_docente(request)
    if not docente:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    estudiante_id = request.POST.get('estudiante_id')
    curso_id      = request.POST.get('curso_id')
    materia_id    = request.POST.get('materia_id')

    asignacion = get_object_or_404(
        AsignacionDocente,
        docente    = docente,
        curso_id   = curso_id,
        materia_id = materia_id,
    )
    Calificacion.objects.filter(
        estudiante_id = estudiante_id,
        asignacion    = asignacion,
    ).delete()

    return JsonResponse({'ok': True})


# ─────────────────────────────────────────
# REPORTES
# ─────────────────────────────────────────

@login_required
def reporte_notas(request):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    cursos   = Curso.objects.filter(asignaciones__docente=docente).distinct()
    materias = Materia.objects.filter(asignaciones__docente=docente).distinct()

    curso_sel   = None
    materia_sel = None
    calificaciones = []
    promedio_general = 0

    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')

    if curso_id and materia_id:
        curso_sel   = get_object_or_404(Curso,   id=curso_id)
        materia_sel = get_object_or_404(Materia, id=materia_id)

        asignacion = AsignacionDocente.objects.filter(
            docente    = docente,
            curso      = curso_sel,
            materia    = materia_sel,
        ).first()

        if asignacion:
            calificaciones = (
                Calificacion.objects
                .filter(asignacion=asignacion)
                .select_related('estudiante')
            )
            promedios = [c.promedio for c in calificaciones]
            promedio_general = round(sum(promedios) / len(promedios), 2) if promedios else 0

    context = {
        'docente':           docente,
        'cursos':            cursos,
        'materias':          materias,
        'curso_sel':         curso_sel,
        'materia_sel':       materia_sel,
        'calificaciones':    calificaciones,
        'promedio_general':  promedio_general,
    }
    return render(request, 'docente/reporte_notas.html', context)


@login_required
def exportar_reporte_pdf(request):
    """Exporta el reporte de notas como PDF usando reportlab."""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter

    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')

    if not (curso_id and materia_id):
        return redirect('reporte_notas')

    asignacion = get_object_or_404(
        AsignacionDocente,
        docente    = docente,
        curso_id   = curso_id,
        materia_id = materia_id,
    )
    calificaciones = (
        Calificacion.objects
        .filter(asignacion=asignacion)
        .select_related('estudiante')
    )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_{asignacion.curso}_{asignacion.materia}.pdf"'
    )

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Encabezado
    p.setFont('Helvetica-Bold', 16)
    p.drawString(50, height - 60, f'Reporte de Notas — {asignacion.materia} — {asignacion.curso}')
    p.setFont('Helvetica', 11)
    p.drawString(50, height - 80, f'Docente: {docente.nombre}')

    # Tabla
    y = height - 120
    p.setFont('Helvetica-Bold', 11)
    p.drawString(50,  y, 'Estudiante')
    p.drawString(260, y, 'Tarea')
    p.drawString(340, y, 'Parcial')
    p.drawString(420, y, 'Examen')
    p.drawString(500, y, 'Promedio')
    y -= 20

    p.setFont('Helvetica', 10)
    for cal in calificaciones:
        if y < 60:
            p.showPage()
            y = height - 60
        p.drawString(50,  y, cal.estudiante.nombre[:35])
        p.drawString(260, y, str(cal.tarea))
        p.drawString(340, y, str(cal.parcial))
        p.drawString(420, y, str(cal.examen))
        p.drawString(500, y, str(cal.promedio))
        y -= 18

    p.showPage()
    p.save()
    return response


@login_required
def exportar_reporte_excel(request):
    """Exporta el reporte de notas como Excel usando openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, Alignment

    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')

    if not (curso_id and materia_id):
        return redirect('reporte_notas')

    asignacion = get_object_or_404(
        AsignacionDocente,
        docente    = docente,
        curso_id   = curso_id,
        materia_id = materia_id,
    )
    calificaciones = (
        Calificacion.objects
        .filter(asignacion=asignacion)
        .select_related('estudiante')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    # Encabezado
    ws.merge_cells('A1:E1')
    ws['A1'] = f'Reporte de Notas — {asignacion.materia} — {asignacion.curso}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append(['Estudiante', 'Tarea', 'Parcial', 'Examen', 'Promedio'])
    for cell in ws[2]:
        cell.font = Font(bold=True)

    for cal in calificaciones:
        ws.append([
            cal.estudiante.nombre,
            float(cal.tarea),
            float(cal.parcial),
            float(cal.examen),
            float(cal.promedio),
        ])

    # Ajuste de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_{asignacion.curso}_{asignacion.materia}.xlsx"'
    )
    wb.save(response)
    return response


# ─────────────────────────────────────────
# ASISTENCIA
# ─────────────────────────────────────────

@login_required
def asistencia_docente(request):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    # Cursos del docente para el selector
    cursos = Curso.objects.filter(asignaciones__docente=docente).distinct()

    curso_sel  = None
    estudiantes = []
    fecha_hoy  = timezone.now().date()

    # Asistencias ya registradas hoy para este curso (si se seleccionó)
    asistencias_hoy = {}  # {estudiante.id: estado}

    curso_id = request.GET.get('curso')
    if curso_id:
        curso_sel   = get_object_or_404(Curso, id=curso_id)
        estudiantes = Estudiante.objects.filter(curso=curso_sel)
        asis = Asistencia.objects.filter(
            estudiante__in=estudiantes,
            fecha=fecha_hoy,
        )
        asistencias_hoy = {a.estudiante_id: a.estado for a in asis}

    context = {
        'docente':         docente,
        'cursos':          cursos,
        'curso_sel':       curso_sel,
        'estudiantes':     estudiantes,
        'asistencias_hoy': asistencias_hoy,
        'fecha_hoy':       fecha_hoy,
    }
    return render(request, 'docente/asistencia_docente.html', context)


@login_required
@require_POST
def guardar_asistencia(request):
    docente = _get_docente(request)

    if not docente:
        return JsonResponse(
            {'error': 'No autorizado'},
            status=403
        )

    try:
        estudiante_id = request.POST.get('estudiante_id')
        fecha = request.POST.get('fecha')
        estado = request.POST.get('estado', 'presente')

        if not estudiante_id:
            return JsonResponse(
                {'error': 'Falta estudiante_id'},
                status=400
            )

        if not fecha:
            return JsonResponse(
                {'error': 'Falta fecha'},
                status=400
            )

        estudiante = get_object_or_404(
            Estudiante,
            id=estudiante_id
        )

        Asistencia.objects.update_or_create(
            estudiante=estudiante,
            fecha=fecha,
            defaults={
                'estado': estado
            }
        )

        return JsonResponse({
            'ok': True
        })

    except Exception as e:
        return JsonResponse(
            {'error': str(e)},
            status=400
        )


@login_required
def historial_asistencia(request):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    cursos = Curso.objects.filter(asignaciones__docente=docente).distinct()

    curso_sel  = None
    fecha_sel  = None
    filas      = []

    curso_id   = request.GET.get('curso')
    fecha_str  = request.GET.get('fecha')   # el template manda "fecha", no "fecha_inicio"

    if curso_id and fecha_str:
        curso_sel = get_object_or_404(Curso, id=curso_id)
        fecha_sel = fecha_str  # 'YYYY-MM-DD', Django lo compara bien contra un DateField

        estudiantes = Estudiante.objects.filter(curso=curso_sel)

        asistencias = Asistencia.objects.filter(
            estudiante__in=estudiantes,
            fecha=fecha_sel,
        )
        asistencia_map = {a.estudiante_id: a for a in asistencias}

        filas = [
            {'estudiante': est, 'asistencia': asistencia_map.get(est.id)}
            for est in estudiantes
        ]

    context = {
        'docente':    docente,
        'cursos':     cursos,
        'curso_sel':  curso_sel,
        'fecha_sel':  fecha_sel,
        'filas':      filas,
    }
    return render(request, 'docente/historial_asistencia.html', context)

@login_required
@require_POST
def eliminar_asistencia(request):
    docente = _get_docente(request)
    if not docente:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    asistencia_id = request.POST.get('asistencia_id')
    asistencia    = get_object_or_404(Asistencia, id=asistencia_id)

    # Verificamos que el estudiante pertenezca a un curso del docente
    es_del_docente = AsignacionDocente.objects.filter(
        docente=docente, curso=asistencia.estudiante.curso
    ).exists()

    if not es_del_docente:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    asistencia.delete()
    return JsonResponse({'ok': True})


# ─────────────────────────────────────────
# MENSAJES
# ─────────────────────────────────────────

@login_required
def mensajes_docente(request):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    # Estudiantes de los cursos del docente
    cursos_docente = Curso.objects.filter(asignaciones__docente=docente).distinct()
    estudiantes    = Estudiante.objects.filter(curso__in=cursos_docente)

    # Conversación seleccionada
    est_sel      = None
    conversacion = []

    estudiante_id = request.GET.get('estudiante')
    if estudiante_id:
        est_sel = get_object_or_404(Estudiante, id=estudiante_id)
        conversacion = Mensaje.objects.filter(
            docente=docente, estudiante=est_sel
        ).order_by('fecha')

        # Marcar como leídos los mensajes del estudiante
        Mensaje.objects.filter(
            docente=docente, estudiante=est_sel,
            enviado_por='estudiante', leido=False,
        ).update(leido=True)

    if request.method == 'POST':
        contenido     = request.POST.get('contenido', '').strip()
        estudiante_id = request.POST.get('estudiante_id')
        if contenido and estudiante_id:
            est = get_object_or_404(Estudiante, id=estudiante_id)
            Mensaje.objects.create(
                docente    = docente,
                estudiante = est,
                contenido  = contenido,
                enviado_por = 'docente',
            )
        return redirect(f"{request.path}?estudiante={estudiante_id}")

    context = {
        'docente':      docente,
        'estudiantes':  estudiantes,
        'est_sel':      est_sel,
        'conversacion': conversacion,
    }
    return render(request, 'docente/mensajes_docente.html', context)


# ─────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────

@login_required
def configuracion_docente(request):
    docente = _get_docente(request)
    if not docente:
        return redirect('login')

    if request.method == 'POST':
        docente.nombre   = request.POST.get('nombre',   docente.nombre).strip()
        docente.correo   = request.POST.get('correo',   docente.correo).strip()
        docente.telefono = request.POST.get('telefono', docente.telefono).strip()
        docente.save()
        return redirect('configuracion_docente')

    context = {'docente': docente}
    return render(request, 'docente/configuracion_docente.html', context)